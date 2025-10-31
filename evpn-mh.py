#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EVPN MULTIHOMING Fabric Builder  - Version 0.1
=== USE AT YOUR OWN RISK! ===
- Reads Excel (FABRIC, INTERFACES, NETWORKS).
- Builds EVPN multihoming topology and per-device configs.
- Detects existing topologies by name and UPDATES them instead of creating new ones
- Uses hostnames instead of device IDs; resolves via Mist API.
- Corrects site settings 'networks' (includes gateway/gateway6).
- Assigns per-core other_ip_configs using gateway+1 and gateway+2
  for the first two collapsed-core switches (name-agnostic).
- Uses STATIC_ROUTESv4 and STATIC_ROUTESv6 with format: route@nexthop
  Supports multiple routes separated by spaces.
- UPDATE MODE: Preserves user-configured ports during fabric updates
- Caches device configs before topology update to intelligently merge port configs
- SPEED/CHANNELIZATION: Configures port speed (10G, 25G, 50G, 100G, 200G, AUTO)
  and channelization (TRUE/FALSE) from INTERFACES sheet

Author: Lukas Eisenberger (leisenberger@juniper.net)

"""

import sys
import ipaddress
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional

from openpyxl import load_workbook

# Add parent directory to path for mistClient/mistHelpers
sys.path.insert(0, str(Path(__file__).parent.parent))

import mistClient
import mistHelpers

# =============================================================================
# CONFIG:
# =============================================================================
MIST_ORGID = "MIST-ORG-ID"
MIST_TOKEN = "API-TOKEN"
MIST_API_URL = "https://api.eu.mist.com"
spreadsheetname = "evpn-mh.xlsx"


# =============================================================================
# ------------------------ helpers ------------------------
def _b(v):
    if isinstance(v, bool): return v
    if v is None: return False
    s = str(v).strip().lower()
    return s in ("1", "true", "t", "y", "yes", "on")


def _i(v, default=None):
    try:
        if v is None or v == "":
            return default
        return int(v)
    except Exception:
        try:
            return int(float(v))
        except Exception:
            return default


def _parse_static_routes(routes_str: Optional[str]) -> Dict[str, Dict[str, str]]:
    """
    Parse static routes in format: route@nexthop route@nexthop ...
    Examples:
      "0.0.0.0/0@192.168.10.254"
      "192.168.66.0/24@192.168.10.254 192.168.77.0/24@192.168.10.254"
      "::/0@2001:10::254"
    Returns dict: {route: {"via": nexthop}, ...}
    """
    result: Dict[str, Dict[str, str]] = {}
    if not routes_str:
        return result

    try:
        # Split by whitespace to get individual route@nexthop pairs
        routes_str = str(routes_str).strip()
        if not routes_str:
            return result

        pairs = routes_str.split()
        for pair in pairs:
            pair = pair.strip()
            if not pair or '@' not in pair:
                continue

            parts = pair.split('@', 1)
            if len(parts) != 2:
                continue

            route = parts[0].strip()
            nexthop = parts[1].strip()

            if not route or not nexthop:
                continue

            # Validate it's a valid network (this will raise if invalid)
            ipaddress.ip_network(route, strict=False)

            # Store in result
            result[route] = {"via": nexthop}
    except Exception as e:
        print(f"Warning: Error parsing static routes '{routes_str}': {e}")
        pass

    return result


def _parse_fabric(wb) -> Dict[str, Any]:
    if "FABRIC" not in wb.sheetnames:
        raise Exception("FABRIC sheet missing")
    sh = wb["FABRIC"]
    data = {}
    for row in sh.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]: continue
        key = str(row[0]).strip()
        val = row[1]
        data[key] = val
    # normalize
    data.setdefault("topologyname", data.get("topology_name"))
    data.setdefault("site_name", data.get("sitename") or data.get("site"))
    data["use_ipv6_underlay"] = _b(data.get("use_ipv6_underlay"))
    data["per_vlan_vga_v4_mac"] = _b(data.get("per_vlan_vga_v4_mac"))
    data["per_vlan_vga_v6_mac"] = _b(data.get("per_vlan_vga_v6_mac"))
    if data.get("overlay_as") is not None:
        data["overlay_as"] = _i(data["overlay_as"])
    if data.get("base_as") is not None:
        data["base_as"] = _i(data["base_as"])
    data["esi_lag_name"] = str(data.get("esi_lag_name") or "EVPN-ESI-LAG")
    return data


def _parse_interfaces(wb) -> List[Dict[str, Any]]:
    """
    Parse INTERFACES sheet with speed and channelization support

    Expected columns: SRC_DEVICE, SRC_DEVICE_ROLE, SRC_INT, DST_DEVICE, DST_DEVICE_ROLE, DST_INT, AE_IDX, SPEED, CHANNELIZED

    SPEED: 10G, 25G, 50G, 100G, 200G, AUTO (case-insensitive)
    CHANNELIZED: TRUE/FALSE (boolean)

    Returns list of interface connection dicts
    """
    if "INTERFACES" not in wb.sheetnames:
        raise Exception("INTERFACES sheet missing")
    sh = wb["INTERFACES"]

    # Get headers from row 1
    header_row = next(sh.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip().upper() if h else "" for h in header_row]

    # Find column indices
    def find_col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    src_dev_idx = find_col("SRC_DEVICE")
    src_role_idx = find_col("SRC_DEVICE_ROLE")
    src_int_idx = find_col("SRC_INT")
    dst_dev_idx = find_col("DST_DEVICE")
    dst_role_idx = find_col("DST_DEVICE_ROLE")
    dst_int_idx = find_col("DST_INT")
    ae_idx_idx = find_col("AE_IDX")
    speed_idx = find_col("SPEED")
    channelized_idx = find_col("CHANNELIZED")

    if src_dev_idx is None or src_int_idx is None or dst_dev_idx is None or dst_int_idx is None:
        raise Exception("INTERFACES sheet missing required columns (SRC_DEVICE, SRC_INT, DST_DEVICE, DST_INT)")

    out = []
    for row in sh.iter_rows(min_row=2, values_only=True):
        if not row or not row[src_dev_idx]:
            continue

        # Parse basic interface data
        src_dev = str(row[src_dev_idx]).strip() if row[src_dev_idx] else None
        src_role = str(row[src_role_idx]).strip() if src_role_idx is not None and len(row) > src_role_idx and row[
            src_role_idx] else None
        src_int = str(row[src_int_idx]).strip() if row[src_int_idx] else None
        dst_dev = str(row[dst_dev_idx]).strip() if row[dst_dev_idx] else None
        dst_role = str(row[dst_role_idx]).strip() if dst_role_idx is not None and len(row) > dst_role_idx and row[
            dst_role_idx] else None
        dst_int = str(row[dst_int_idx]).strip() if row[dst_int_idx] else None
        ae_idx = _i(row[ae_idx_idx]) if ae_idx_idx is not None and len(row) > ae_idx_idx and row[ae_idx_idx] else None

        # Parse speed (optional)
        speed = None
        if speed_idx is not None and len(row) > speed_idx and row[speed_idx]:
            speed_raw = str(row[speed_idx]).strip().upper()
            # Normalize speed format: 10G, 25G, 50G, 100G, 200G, AUTO
            if speed_raw in ["10G", "25G", "50G", "100G", "200G", "AUTO"]:
                speed = speed_raw.lower()  # Mist API uses lowercase
            else:
                # Try to extract just the number+G part
                import re
                match = re.match(r'(\d+)G?', speed_raw, re.IGNORECASE)
                if match:
                    speed = f"{match.group(1)}g"
                elif speed_raw == "AUTO":
                    speed = "auto"

        # Parse channelized (optional)
        channelized = None
        if channelized_idx is not None and len(row) > channelized_idx and row[channelized_idx]:
            channelized = _b(row[channelized_idx])

        out.append({
            "src_device": src_dev,
            "src_role": src_role,
            "src_int": src_int,
            "dst_device": dst_dev,
            "dst_role": dst_role,
            "dst_int": dst_int,
            "ae_idx": ae_idx,
            "speed": speed,
            "channelized": channelized,
        })

    if not out:
        raise Exception("INTERFACES has no rows")
    return out


def _roles_from_interfaces(entries: List[Dict[str, Any]]) -> Dict[str, str]:
    roles: Dict[str, str] = {}
    for e in entries:
        for dev, role in ((e["src_device"], e["src_role"]), (e["dst_device"], e["dst_role"])):
            if not dev or not role: continue
            if dev in roles and roles[dev] != role:
                raise Exception(f"Device '{dev}' has conflicting roles: '{roles[dev]}' vs '{role}'")
            roles[dev] = role
    return roles


def _parse_network_comment_directives(comment: Optional[str]) -> Tuple[
    Dict[str, Dict[str, str]], Dict[str, Dict[str, str]]]:
    r4: Dict[str, Dict[str, str]] = {}
    r6: Dict[str, Dict[str, str]] = {}
    if not comment: return r4, r6
    try:
        parts = [p.strip() for p in str(comment).split(";") if p.strip()]
        for p in parts:
            if " via " in p:
                prefix, nh = [x.strip() for x in p.split(" via ", 1)]
                if prefix.lower().startswith("default"): prefix = "0.0.0.0/0"
                if ":" in prefix:
                    ipaddress.ip_network(prefix, strict=False)
                    r6[prefix] = {"via": nh}
                else:
                    ipaddress.ip_network(prefix, strict=False)
                    r4[prefix] = {"via": nh}
    except Exception:
        pass
    return r4, r6


def _parse_networks(wb) -> Tuple[Dict[str, Any], Dict[str, Any], List[str], List[Dict[str, Any]]]:
    if "NETWORKS" not in wb.sheetnames:
        raise Exception("NETWORKS sheet missing")
    sh = wb["NETWORKS"]
    header = [str(c).strip() if c else "" for c in next(sh.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []
    for row in sh.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None: continue
        rows.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
    site_networks: Dict[str, Any] = {}
    vrf_instances: Dict[str, Any] = {}
    network_name_list: List[str] = []
    extra_v4: Dict[str, Dict[str, Dict[str, str]]] = {}
    extra_v6: Dict[str, Dict[str, Dict[str, str]]] = {}
    for net in rows:
        name = str(net.get("NETWORKNAME") or "").strip()
        if not name: continue
        network_name_list.append(name)
        vlan_raw = net.get("VLAN_ID")
        vlan_id = _i(vlan_raw, default=vlan_raw)
        try:
            vlan_id = int(vlan_id)
        except Exception:
            pass  # keep as-is if text

        vrf = str(net.get("VRF") or "").strip()
        gw4 = net.get("GATEWAY")
        gw6 = net.get("GATEWAY6")

        # NEW: Parse STATIC_ROUTESv4 and STATIC_ROUTESv6
        static_routes_v4_str = net.get("STATIC_ROUTESv4")
        static_routes_v6_str = net.get("STATIC_ROUTESv6")

        comment = net.get("COMMENT")

        subnet4 = ""
        subnet6 = ""
        gw4_ip = ""
        gw6_ip = ""
        if gw4:
            i4 = ipaddress.ip_interface(str(gw4))
            subnet4 = f"{i4.network.network_address}/{i4.network.prefixlen}"
            gw4_ip = str(i4.ip)
        if gw6:
            i6 = ipaddress.ip_interface(str(gw6))
            subnet6 = f"{i6.network.network_address}/{i6.network.prefixlen}"
            gw6_ip = str(i6.ip)

        site_networks[name] = {
            "vlan_id": vlan_id,
            "subnet": subnet4,
            "subnet6": subnet6,
        }
        if gw4_ip: site_networks[name]["gateway"] = gw4_ip
        if gw6_ip: site_networks[name]["gateway6"] = gw6_ip

        if vrf:
            inst = vrf_instances.setdefault(vrf, {"networks": [], "v4_routing": False, "v6_routing": False})
            inst["networks"].append(name)
            if gw4_ip: inst["v4_routing"] = True
            if gw6_ip: inst["v6_routing"] = True

            # Parse static routes from new format
            if static_routes_v4_str:
                routes_v4 = _parse_static_routes(static_routes_v4_str)
                if routes_v4:
                    extra_v4.setdefault(vrf, {}).update(routes_v4)

            if static_routes_v6_str:
                routes_v6 = _parse_static_routes(static_routes_v6_str)
                if routes_v6:
                    extra_v6.setdefault(vrf, {}).update(routes_v6)

            # Also parse comment directives (for backward compatibility)
            r4, r6 = _parse_network_comment_directives(comment)
            if r4: extra_v4.setdefault(vrf, {}).update(r4)
            if r6: extra_v6.setdefault(vrf, {}).update(r6)

    for vrf, inst in vrf_instances.items():
        if vrf in extra_v4: inst["extra_routes"] = extra_v4[vrf]
        if vrf in extra_v6: inst["extra_routes6"] = extra_v6[vrf]
    return site_networks, vrf_instances, network_name_list, rows


def _find_existing_topology(mist: mistClient.Mist, site_id: str, topology_name: str) -> Optional[Dict[str, Any]]:
    """
    Search for an existing EVPN topology by name in the specified site
    Returns the topology dict if found, None otherwise
    """
    try:
        topos = mist.get(f"sites/{site_id}/evpn_topologies")
        if not topos:
            return None

        for topo in topos:
            if topo.get("name") == topology_name:
                return topo

        return None
    except Exception as e:
        print(f"Warning: Failed to retrieve existing topologies: {e}")
        return None


def _get_device_config(mist: mistClient.Mist, site_id: str, device_id: str) -> Dict[str, Any]:
    """
    Fetch current device configuration from Mist API
    Returns the complete device configuration dict
    """
    try:
        url = f"sites/{site_id}/devices/{device_id}"
        device = mist.get(url)
        return device if device else {}
    except Exception as e:
        print(f"Warning: Failed to retrieve device config for {device_id}: {e}")
        return {}


def _merge_port_configs(
        current_port_config: Dict[str, Any],
        new_evpn_ports: Dict[str, Any],
        esi_lag_name: str
) -> Dict[str, Any]:
    """
    Intelligently merge port configurations, preserving user-configured ports.

    Strategy:
    1. Start with current port config (preserves everything)
    2. Update/add EVPN ports (uplink/downlink, ESI-LAG)

    This ensures that user-configured ports (like access ports, trunk ports, etc.)
    are never touched by the script.

    Args:
        current_port_config: Current port config from device (CACHED from before topology update)
        new_evpn_ports: New EVPN uplink/downlink/ESI-LAG ports from Excel
        esi_lag_name: Name of ESI-LAG port usage (e.g., "EVPN-ESI-LAG")

    Returns:
        Merged port configuration
    """
    merged_config = {}

    # Start with current config (preserves all existing user-configured ports)
    for port_name, port_cfg in current_port_config.items():
        merged_config[port_name] = port_cfg

    # Update with new EVPN ports (these take precedence)
    for port_name, port_cfg in new_evpn_ports.items():
        merged_config[port_name] = port_cfg

    return merged_config


def _build_optic_port_config(entries: List[Dict[str, Any]], device_roles: Dict[str, str]) -> Dict[str, Dict[str, Any]]:
    """
    Build optic_port_config from interface entries with speed/channelization settings

    Args:
        entries: List of interface connection dicts (from _parse_interfaces)
        device_roles: Dict mapping device names to roles

    Returns:
        Dict[device_name, Dict[base_port, {speed, channelized}]]

    Example output:
        {
            "D1": {
                "et-0/0/0": {"speed": "10g", "channelized": True},
                "et-0/0/16": {"speed": "100g", "channelized": False}
            }
        }
    """
    optic_config: Dict[str, Dict[str, Any]] = {}

    # Process each interface entry
    for e in entries:
        speed = e.get("speed")
        channelized = e.get("channelized")

        # Skip if no speed/channelization specified
        if speed is None and channelized is None:
            continue

        # Extract base port name (remove :0, :1, :2, :3 suffix for channelized ports)
        def get_base_port(port_name: str) -> str:
            """Extract base port name, removing channelization suffix (:0, :1, etc.)"""
            if not port_name:
                return port_name
            # Remove channelization suffix like :0, :1, :2, :3
            import re
            match = re.match(r'^([^:]+)(?::\d+)?$', port_name)
            if match:
                return match.group(1)
            return port_name

        # Process source device/interface
        if e.get("src_device") and e.get("src_int"):
            src_dev = e["src_device"]
            src_base_port = get_base_port(e["src_int"])

            if src_dev not in optic_config:
                optic_config[src_dev] = {}

            # Only add if not already configured (first occurrence wins)
            if src_base_port not in optic_config[src_dev]:
                config_entry = {}
                if speed is not None:
                    config_entry["speed"] = speed
                if channelized is not None:
                    config_entry["channelized"] = channelized

                if config_entry:  # Only add if we have something to configure
                    optic_config[src_dev][src_base_port] = config_entry

        # Process destination device/interface
        if e.get("dst_device") and e.get("dst_int"):
            dst_dev = e["dst_device"]
            dst_base_port = get_base_port(e["dst_int"])

            if dst_dev not in optic_config:
                optic_config[dst_dev] = {}

            # Only add if not already configured (first occurrence wins)
            if dst_base_port not in optic_config[dst_dev]:
                config_entry = {}
                if speed is not None:
                    config_entry["speed"] = speed
                if channelized is not None:
                    config_entry["channelized"] = channelized

                if config_entry:  # Only add if we have something to configure
                    optic_config[dst_dev][dst_base_port] = config_entry

    return optic_config


def _build_topology(entries: List[Dict[str, Any]], device_roles: Dict[str, str]) -> Tuple[
    List[Dict[str, str]], Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    core_devices = [d for d, r in device_roles.items() if r == "collapsed-core"]
    if len(core_devices) < 2:
        raise Exception("Need at least two devices with role 'collapsed-core'")
    # stable order
    core_devices.sort()
    # core-core links
    core_links: List[Dict[str, str]] = []
    seen = set()
    for e in entries:
        if e["src_device"] in core_devices and e["dst_device"] in core_devices:
            a = (e["src_device"], e["src_int"])
            b = (e["dst_device"], e["dst_int"])
            key = frozenset({a, b})
            if key in seen: continue
            seen.add(key)
            core_links.append(
                {"core1": e["src_device"], "port1": e["src_int"], "core2": e["dst_device"], "port2": e["dst_int"]})

    access_port_config: Dict[str, Dict[str, Any]] = {}
    core_port_config: Dict[str, Dict[str, Any]] = {c: {} for c in core_devices}
    ae_groups: Dict[int, List[Dict[str, Any]]] = {}
    for e in entries:
        if e["ae_idx"] is None: continue
        s_is_core = device_roles.get(e["src_device"]) == "collapsed-core"
        d_is_core = device_roles.get(e["dst_device"]) == "collapsed-core"
        if s_is_core ^ d_is_core:
            ae_groups.setdefault(e["ae_idx"], []).append(e)

    for ae, conns in ae_groups.items():
        if len(conns) != 2:
            raise Exception(f"AE_IDX={ae} must have exactly 2 connections (found {len(conns)})")
        access_dev = None
        access_ports: List[str] = []
        for c in conns:
            if device_roles.get(c["src_device"]) == "collapsed-core":
                core_dev, core_port = c["src_device"], c["src_int"]
                access_dev, access_port = c["dst_device"], c["dst_int"]
            else:
                core_dev, core_port = c["dst_device"], c["dst_int"]
                access_dev, access_port = c["src_device"], c["src_int"]
            if core_port in core_port_config[core_dev]:
                raise Exception(f"Core '{core_dev}' port '{core_port}' used in multiple AE groups")
            core_port_config[core_dev][core_port] = {"usage": "EVPN-ESI-LAG", "aggregated": True, "esilag": True,
                                                     "ae_idx": ae}
            access_ports.append(access_port)
        port_key = ",".join(sorted({p for p in access_ports if p}))
        access_port_config[access_dev] = {port_key: {"usage": "EVPN-ESI-LAG", "aggregated": True, "ae_idx": ae}}

    for i, l in enumerate(sorted(core_links, key=lambda x: (x["core1"], x["core2"], x["port1"], x["port2"])), start=1):
        c1, p1, c2, p2 = l["core1"], l["port1"], l["core2"], l["port2"]
        if i % 2 == 1:
            core_port_config[c1][p1] = {"usage": "evpn_downlink", "link_name": f"link{i}"}
            core_port_config[c2][p2] = {"usage": "evpn_uplink", "link_name": f"link{i}"}
        else:
            core_port_config[c1][p1] = {"usage": "evpn_uplink", "link_name": f"link{i}"}
            core_port_config[c2][p2] = {"usage": "evpn_downlink", "link_name": f"link{i}"}
    return core_links, core_port_config, access_port_config


def create_fabric(xlsx_path: str):
    print("Connecting to Mist API...")
    mist = mistClient.Mist(MIST_API_URL, MIST_TOKEN, MIST_ORGID)
    mh = mistHelpers.MistHelpers(mist)
    if not mist.test_connection():
        raise Exception("Failed to connect to Mist API")
    print("Successfully connected to API.\n")

    wb = load_workbook(xlsx_path)

    fabric = _parse_fabric(wb)
    site_name = fabric.get("site_name") or fabric.get("topologyname")
    topo_name = fabric.get("topologyname")
    if not site_name or not topo_name:
        raise Exception("FABRIC needs 'site_name' and 'topologyname'")

    # Lookup site id
    site_id = None
    for s in mh.get_sites():
        if s.get("name") == site_name:
            site_id = s.get("id")
            break
    if not site_id:
        raise Exception(f"Site '{site_name}' not found")

    print(f"Found site: {site_name} (ID: {site_id})")

    # =========================================================================
    # CHECK FOR EXISTING TOPOLOGY
    # =========================================================================
    existing_topology = _find_existing_topology(mist, site_id, topo_name)

    if existing_topology:
        print(f"\n*** EXISTING EVPN TOPOLOGY FOUND ***")
        print(f"EVPN Topology '{topo_name}' already exists with ID: {existing_topology['id']}")
        print(f"The EVPN Topology will be UPDATED according to the Excel configuration")
        mode = "UPDATE"
        topo_id = existing_topology["id"]
    else:
        print(f"\n*** NO EXISTING EVPN TOPOLOGY FOUND ***")
        print(f"EVPN Topology '{topo_name}' will be CREATED from scratch")
        mode = "CREATE"
        topo_id = None

    # devices in site
    devs = mh.get_switches(site_id)
    dev_by_name = {d.get("name"): d for d in devs}

    # parse excel
    entries = _parse_interfaces(wb)
    roles = _roles_from_interfaces(entries)
    core_devices = sorted([d for d, r in roles.items() if r == "collapsed-core"])
    if len(core_devices) < 2:
        raise Exception("Need two collapsed-core devices in INTERFACES")

    # verify devices exist
    for name in roles.keys():
        if name not in dev_by_name:
            raise Exception(f"Device '{name}' not found in site '{site_name}'")

    site_networks, vrf_instances, network_name_list, networks_rows = _parse_networks(wb)

    # build port configs
    core_links, core_port_config, access_port_config = _build_topology(entries, roles)

    # build optic port configuration from speed/channelization settings
    optic_port_config = _build_optic_port_config(entries, roles)
    print(f"\nBuilt optic_port_config for {len(optic_port_config)} devices")
    for dev_name, ports in optic_port_config.items():
        print(f"  {dev_name}: {len(ports)} ports configured")

    # =========================================================================
    # CACHE CURRENT DEVICE CONFIGS (for UPDATE mode)
    # =========================================================================
    # Cache device configs BEFORE updating topology
    # This preserves user-configured ports during updates
    device_configs_cache = {}
    if mode == "UPDATE":
        print(f"\n=== Caching current device configurations ===")
        for name in roles.keys():
            dev_id = dev_by_name[name].get("id")
            device_configs_cache[name] = _get_device_config(mist, site_id, dev_id)
            print(f"Cached config for: {name}")
        print("Device configs cached successfully")

    # 1) Site settings with networks + vrf + EVPN-ESI-LAG usage (no topo id yet)
    print("\n=== STEP 1: Updating site settings ===")
    print("Site settings payload:")
    print(f"  - VRF instances: {list(vrf_instances.keys())}")
    print(f"  - Networks: {len(site_networks)}")
    print(f"  - Port usage: {fabric['esi_lag_name']}")

    site_setting_payload = {
        "vrf_instances": vrf_instances,
        "networks": site_networks,
        "port_usages": {
            fabric["esi_lag_name"]: {
                "mode": "trunk",
                "disabled": False,
                "port_network": None,
                "voip_network": None,
                "stp_edge": False,
                "all_networks": False,
                "networks": network_name_list,
                "port_auth": None,
                "speed": "auto",
                "duplex": "auto",
                "mac_limit": "0",
                "poe_disabled": True,
                "enable_qos": False,
                "storm_control": {},
                "mtu": "9200"
            }
        }
    }
    resp = mist.put(f"sites/{site_id}/setting", site_setting_payload)
    print("Site settings updated successfully")

    # 2) Create or Update EVPN topology
    print(f"\n=== STEP 2: {'Updating' if mode == 'UPDATE' else 'Creating'} EVPN topology ===")

    evpn_options = {
        "routed_at": "edge",
        "overlay": {"as": fabric.get("overlay_as")},
        "core_as_border": True,
        "per_vlan_vga_v4_mac": fabric.get("per_vlan_vga_v4_mac", False),
        "per_vlan_vga_v6_mac": fabric.get("per_vlan_vga_v6_mac", False),
        "underlay": {
            "as_base": fabric.get("base_as"),
            "use_ipv6": fabric.get("use_ipv6_underlay", False),
            "subnet": fabric.get("underlay_subnet"),
        }
    }
    if fabric.get("use_ipv6_underlay"):
        if fabric.get("auto_router_id_subnet6"):
            evpn_options["auto_router_id_subnet6"] = fabric.get("auto_router_id_subnet6")
    if fabric.get("auto_router_id_subnet"):
        evpn_options["auto_router_id_subnet"] = fabric.get("auto_router_id_subnet")

    switches = []
    # stable core order; first two are the cores that get +1/+2
    for name, role in roles.items():
        d = dev_by_name[name]
        mac = (d.get("mac") or "").replace(":", "").replace("-", "").replace(".", "").lower()
        e = {"mac": mac, "role": role, "uplinks": [], "downlinks": [], "config": {"port_config": {}}}
        if role == "collapsed-core":
            # link to the other core
            other = [c for c in core_devices if c != name]
            if other:
                omac = (dev_by_name[other[0]].get("mac") or "").replace(":", "").replace("-", "").replace(".",
                                                                                                          "").lower()
                if omac: e["uplinks"] = [omac]; e["downlinks"] = [omac]
            e["config"]["port_config"] = core_port_config.get(name, {})
        else:
            e["config"]["port_config"] = access_port_config.get(name, {})
        switches.append(e)

    topo_payload = {"name": topo_name, "overwrite": True, "evpn_options": evpn_options, "pod_names": {"1": "Pod 1"},
                    "switches": switches}

    if mode == "UPDATE":
        # Use PUT to update existing topology
        print(f"Updating existing topology '{topo_name}' (ID: {topo_id}) with {len(switches)} switches")
        topo_resp = mist.put(f"sites/{site_id}/evpn_topologies/{topo_id}", topo_payload)
        print(f"EVPN topology updated successfully")
    else:
        # Use POST to create new topology
        print(f"Creating new topology '{topo_name}' with {len(switches)} switches")
        topo_resp = mist.post(f"sites/{site_id}/evpn_topologies", topo_payload)
        topo_id = topo_resp.get("id")
        print(f"EVPN topology created with ID: {topo_id}")

    # 3) Per-device updates: build other_ip_configs for cores (+1, +2)
    print("\n=== STEP 3: Configuring individual devices ===")
    # offsets: +1 for first core, +2 for second core (sorted list)
    core_offsets = {core_devices[0]: 1, core_devices[1]: 2}

    def make_other_ip_configs(offset: int) -> Dict[str, Any]:
        cfg = {}
        for net in networks_rows:
            name = str(net.get("NETWORKNAME") or "").strip()
            if not name: continue
            entry = {}
            gw4 = net.get("GATEWAY")
            gw6 = net.get("GATEWAY6")
            if gw4:
                i4 = ipaddress.ip_interface(str(gw4))
                ip4 = i4.ip + offset
                entry.update({"type": "static", "ip": str(ip4), "netmask": str(i4.network.netmask)})
            if gw6:
                i6 = ipaddress.ip_interface(str(gw6))
                ip6 = i6.ip + offset
                entry.update({"type6": "static", "ip6": str(ip6), "netmask6": f"/{i6.network.prefixlen}"})
            if entry:
                cfg[name] = entry
        return cfg

    for name, role in roles.items():
        dev_id = dev_by_name[name].get("id")

        # Get the new EVPN port config from Excel
        new_port_config = core_port_config.get(name, {}) if role == "collapsed-core" else access_port_config.get(name,
                                                                                                                 {})

        # In UPDATE mode, merge with cached config to preserve user-configured ports
        if mode == "UPDATE":
            cached_config = device_configs_cache.get(name, {})
            current_port_config = cached_config.get("port_config", {})
            final_port_config = _merge_port_configs(current_port_config, new_port_config, fabric["esi_lag_name"])
        else:
            final_port_config = new_port_config

        if role == "collapsed-core":
            offset = core_offsets.get(name, 1)  # default first core
            payload = {
                "other_ip_configs": make_other_ip_configs(offset),
                "port_config": final_port_config,
                "optic_port_config": optic_port_config.get(name, {}),
                "vrf_config": {"enabled": bool(vrf_instances)},
                "dhcpd_config": {"enabled": False}
            }
            print(f"Configuring collapsed-core switch: {name}")
        else:
            payload = {
                "other_ip_configs": {},
                "port_config": final_port_config,
                "optic_port_config": optic_port_config.get(name, {}),
                "-ui_evpntopo_id": True
            }
            print(f"Configuring access switch: {name}")
        resp = mist.put(f"sites/{site_id}/devices/{dev_id}", payload)

    # 4) Link EVPN topology to port usage (include networks again to be safe)
    print("\n=== STEP 4: Linking EVPN-ESI-LAG to topology ===")
    final_payload = {
        "port_usages": {
            fabric["esi_lag_name"]: {
                "mode": "trunk",
                "disabled": False,
                "port_network": None,
                "voip_network": None,
                "stp_edge": False,
                "all_networks": False,
                "port_auth": None,
                "speed": "auto",
                "duplex": "auto",
                "mac_limit": "0",
                "poe_disabled": True,
                "enable_qos": False,
                "storm_control": {},
                "mtu": "9200",
                "networks": network_name_list,
                "ui_evpntopo_id": topo_id
            }
        },
        "vrf_instances": vrf_instances,
        "networks": site_networks
    }
    resp = mist.put(f"sites/{site_id}/setting", final_payload)
    print("EVPN-ESI-LAG successfully linked to topology")

    print("\n=== Fabric creation complete ===")
    print(f"Topology: {topo_name}")
    print(f"Topology ID: {topo_id}")
    print(f"Site: {site_name}")
    print(f"Switches configured: {len(roles)}")
    print(f"  - Collapsed-core: {len(core_devices)}")
    print(f"  - Access: {len(roles) - len(core_devices)}")
    print(f"Networks: {len(site_networks)}")
    print(f"VRFs: {len(vrf_instances)}")
    print(f"ESI-LAG groups: {len(set(e['ae_idx'] for e in entries if e.get('ae_idx')))}")
    print(f"Core-to-core links: {len(core_links)}")


def main():
    # Get script directory for finding spreadsheet
    script_dir = Path(__file__).parent

    if len(sys.argv) >= 2:
        xlsx = sys.argv[1]
    else:
        # Look for spreadsheet in same directory as script
        xlsx = str(script_dir / spreadsheetname)

    try:
        create_fabric(xlsx)
    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()